from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import io

def create_excel_with_links(links):
    wb = Workbook()
    ws = wb.active
    ws.title = "İş İlanları"

    # Başlık
    ws['A1'] = "İş İlanı Linkleri"
    ws['A1'].font = Font(bold=True)

    # Linkleri satırlara ekle
    for idx, link in enumerate(links, start=2):
        cell = ws.cell(row=idx, column=1, value=link)
        cell.hyperlink = link
        cell.style = "Hyperlink"

    # Sütun genişliğini ayarla
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

    # Excel dosyasını belleğe yaz
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
